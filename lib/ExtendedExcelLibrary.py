__author__ = 'VUMVO'

from ExcelLibrary import ExcelLibrary
from lib import General
from os import path
import os
from robot.api import logger
from operator import itemgetter
from datetime import datetime, timedelta
from xlrd import open_workbook, cellname, xldate_as_tuple, \
    XL_CELL_NUMBER, XL_CELL_DATE, XL_CELL_TEXT, XL_CELL_BOOLEAN, \
    XL_CELL_ERROR, XL_CELL_BLANK, XL_CELL_EMPTY, error_text_from_code
from xlwt import easyxf, Workbook
from xlutils.copy import copy as copy

from openpyxl import *


class ExtendedExcelLibrary(ExcelLibrary):
    ROBOT_LIBRARY_SCOPE = 'TEST SUITE'

    def __init__(self, file_name, sheetname):
        if file_name[:1] != '\\':
            self.filePath = General.get_absolute_path(file_name[:])
        else:
            self.filePath = file_name
        logger.info('\n' +self.filePath, False, True)

        'Loading the workbook'
        self.sheetName = sheetname
        self.wb = load_workbook(self.filePath, False)
        self.sheet = self.wb.get_sheet_by_name(self.sheetName)

    def get_cell_value_by_column_header(self, row_index, header_name):
        column_header_index = self.get_column_header_index(header_name)
        return self.sheet.cell(row=row_index, column=column_header_index).value

    def get_column_header_index(self, header_name):
        for col_index in range(1, self.sheet.max_column+1):
            value = self.sheet.cell(row=1, column=col_index).value
            logger.console(str(value))
            if str(value) == header_name:
                return col_index

    def get_number_of_rows(self):
        return self.sheet.max_row

    def get_number_of_columns(self):
        return self.sheet.max_column

    def get_row_values(self, row):
        row_values = []
        for col_index in range(1, self.sheet.max_column+1):
            value = self.sheet.cell(row=row, column=col_index).value
            row_values.append(value)
        return row_values

    def get_sheet_values(self):
        sheet_values = []
        for row_index in range(1, self.sheet.max_row+1):
            row_values = []
            for col_index in range(1, self.sheet.max_column+1):
                value = self.sheet.cell(row=row_index, column=col_index).value
                row_values.append(value)
            sheet_values.append(row_values)
        return sheet_values

    def set_cell_value(self, row_index, header_name, value):
        column_index = self.get_column_header_index(header_name)
        logger.console(str(row_index) + ' ' + str(column_index) +'=' + str(value))
        _cell = self.sheet.cell(row=row_index, column=column_index)
        _cell.value = value

    def save_excel_file(self):
        self.wb.save(self.filePath)
